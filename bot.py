import telebot
from telebot import types
import re
import logging
import os
import json
from datetime import datetime
import openpyxl
from io import BytesIO
import time
import subprocess
import base64

# 🔴 Warning মেসেজ বন্ধ করার ম্যাজিক ট্রিক
import warnings
warnings.filterwarnings("ignore")

# 🟢 Groq এর জন্য OpenAI লাইব্রেরি ইম্পোর্ট
from openai import OpenAI

# ════════════════════════════════════════════════════════════════
#  LOCAL MEMORY SETUP (Firebase Removed!)
# ════════════════════════════════════════════════════════════════
local_ai_memory = {}

def save_memory(user_id, text):
    user_id_str = str(user_id)
    old_data = local_ai_memory.get(user_id_str, "")
    updated_data = old_data + "\n" + text
    
    if len(updated_data) > 2000:
        updated_data = updated_data[-2000:]
        
    local_ai_memory[user_id_str] = updated_data

def get_memory(user_id):
    return local_ai_memory.get(str(user_id), "")


# ════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  শুধু এই অংশ পরিবর্তন করুন
# ════════════════════════════════════════════════════════════════
BOT_TOKEN      = os.environ.get("BOT_TOKEN", "8540616234:AAGEtDCYh77B0VQIlEFek57my45S_GtnrS8")
CHANNEL_ID     = -1003869262466
ADMIN_IDS      = [7689218221, 7833093821]
SUPPORT_USER   = "@FBSKYSUPPORT এবং @ONLYALLSUPPORT"
BOT_NAME       = "FB ID Submit Hub"
MAX_HISTORY    = 10

GROQ_API_KEYS = [
    "", 
    "", 
    ""   
]
CURRENT_API_INDEX = 0

CUSTOM_BOT_CONTEXT = ""

SYSTEM_SETTINGS = {
    "min_id_limit": 5,           
    "support_username": SUPPORT_USER, 
    "ai_chat_enabled": True,     
    "maintenance_mode": False,
    "welcome_msg": "📌 *কীভাবে ID জমা দেবেন:*\n1️⃣ *SELL ID* বাটন চাপুন\n2️⃣ ক্যাটাগরি সিলেক্ট করুন\n3️⃣ আপনার Telegram username দিন\n4️⃣ Excel (.xlsx) ফাইল আপলোড করুন\n5️⃣ Payment নম্বর ও নোট দিন\n✅ সাবমিট সম্পন্ন — Admin result জানাবে",
    "notice_board": ""  # 🆕 Notice Board
}

# 🆕 Bot start time for uptime tracking
BOT_START_TIME = datetime.now()

# ════════════════════════════════════════════════════════════════
#  CATEGORIES
# ════════════════════════════════════════════════════════════════
CATEGORIES = {
    "PC1000X":  {"name": "📲 PC Clone 1000x",         "rate": 13.00, "format": "A=UID | B=Password | C=Cookie",              "status": "open"},
    "PC6155X":  {"name": "📲 PC Clone 6155x/56x/57x", "rate":  7.00, "format": "A=UID | B=Password | C=Cookie",              "status": "open"},
    "PC6158X":  {"name": "📲 PC Clone 6158x",         "rate":  4.00, "format": "A=UID | B=Password | C=Cookie",              "status": "open"},
    "NUM2FA":   {"name": "⚡ Number 2FA I'D",          "rate":  6.00, "format": "A=UID | B=Password | C=2FA Key",             "status": "open"},
    "NUMCOOKIE":{"name": "🟢 Number Cookies I'D",      "rate":  4.00, "format": "A=UID | B=Password | C=Cookie",              "status": "open"},
    "HOTMAIL30":{"name": "☁️ Hotmail 30+ Friend",      "rate": 10.00, "format": "A=UID | B=Password | C=2FA | D=Full Mail",   "status": "open"},
    "HOTMAIL00":{"name": "💠 Hotmail 00 Friend",       "rate":  7.00, "format": "A=UID | B=Password | C=2FA | D=Full Mail",   "status": "open"},
    "INSTA2FA": {"name": "📸 Instagram 2FA",           "rate":  2.70, "format": "A=Username | B=Password | C=2FA Key",        "status": "open"},
    "INSTACOOK":{"name": "🍪 Instagram Cookies",       "rate":  4.00, "format": "A=Username | B=Password  (2 columns only)",  "status": "open"},
}

# ════════════════════════════════════════════════════════════════
#  INIT
# ════════════════════════════════════════════════════════════════
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger(__name__)

bot = telebot.TeleBot(BOT_TOKEN, parse_mode=None)

user_data        = {}
all_submissions  = {}
user_submissions = {}
submission_count = 0
registered_users = set()
BANNED_USERS     = set()
username_to_id   = {}
ai_sessions      = {}

# ════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════
def is_admin(cid):        
    return cid in ADMIN_IDS

def is_banned(cid):
    return cid in BANNED_USERS

def safe_md(text):
    escape_chars = r'_*[]()~`>#+-=|{}.!'
    return re.sub(f'([{re.escape(escape_chars)}])', r'\\\1', text)

# 🆕 User Badge/Rank system
def get_user_badge(total_earned):
    if total_earned >= 5000:
        return "💎 Diamond"
    elif total_earned >= 2000:
        return "🥇 Gold"
    elif total_earned >= 500:
        return "🥈 Silver"
    elif total_earned >= 100:
        return "🥉 Bronze"
    else:
        return "🌱 Starter"

MENU_BUTTONS = ["🚀 SELL ID 🚀", "📊 Price List", "🎧 Support", "📜 My History", "👤 My Profile", "🏆 Top Sellers", "🤖 AI Chat", "❌ Stop AI", "🔐 Admin Panel", "🌐 Global Stats", "📩 Feedback", "📣 Notice Board", "🔎 Check Status", "📚 Rules & FAQ"]

def main_menu(cid):
    m = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    m.add(types.KeyboardButton("🚀 SELL ID 🚀"))
    m.row(types.KeyboardButton("📊 Price List"), types.KeyboardButton("🎧 Support"))
    m.row(types.KeyboardButton("👤 My Profile"), types.KeyboardButton("📜 My History"))
    m.row(types.KeyboardButton("🌐 Global Stats"), types.KeyboardButton("📩 Feedback"))
    m.row(types.KeyboardButton("🏆 Top Sellers"), types.KeyboardButton("🤖 AI Chat"))
    m.row(types.KeyboardButton("📣 Notice Board"), types.KeyboardButton("🔎 Check Status"))
    m.row(types.KeyboardButton("📚 Rules & FAQ")) # 🆕 New Rules Button
    if is_admin(cid):
        m.add(types.KeyboardButton("🔐 Admin Panel"))
    m.add(types.KeyboardButton("❌ Stop AI"))
    return m

def cancel_btn():
    m = types.InlineKeyboardMarkup()
    m.add(types.InlineKeyboardButton("❌ CANCEL", callback_data="cancel_flow"))
    return m

def build_price_list():
    lines = [
        "🏷 *Price List*",
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    ]
    for cat in CATEGORIES.values():
        st = "✅ Open" if cat["status"] == "open" else "🔴 Closed"
        lines.append(
            f"\n{'✅' if cat['status']=='open' else '🔴'} *{cat['name']}*\n"
            f"   💰 Rate: *{cat['rate']:.2f} Tk/pcs*\n"
            f"   📋 `{cat['format']}`\n"
            f"   Status: {st}"
        )
    lines.append("\n━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append(f"📞 Support: {SYSTEM_SETTINGS['support_username']}")
    return "\n".join(lines)

def register_user(message):
    registered_users.add(message.chat.id)
    if message.from_user and message.from_user.username:
        username_to_id[f"@{message.from_user.username.lower()}"] = message.chat.id

# ════════════════════════════════════════════════════════════════
#  AI ENGINE — Jarvis (Groq Cloud AI Logic)
# ════════════════════════════════════════════════════════════════
class JarvisAI:
    def __init__(self):
        self.model_name_text = "llama-3.3-70b-versatile" 
        self.model_name_vision = "llama-3.2-11b-vision-preview" 

    def get_response(self, prompt, user_name, is_admin_user=False, image_data=None):
        global CURRENT_API_INDEX, CUSTOM_BOT_CONTEXT

        bot_context = (
            f"You are JARVIS, an AI assistant for the Telegram bot '{BOT_NAME}'. "
            "Your job is to help users sell their IDs, tell them prices, and solve their bot-related problems. "
            "Here is the current price list and categories:\n"
        )
        for code, cat in CATEGORIES.items():
            bot_context += f"- {cat['name']}: {cat['rate']} Tk/pcs (Status: {cat['status']})\n"
        
        if is_admin_user:
            bot_context += "\n[CRITICAL NOTE: The user you are talking to is the OWNER/ADMIN of this bot (Mehedi Boss or Sakhawat Boss). Be highly respectful, assist with system details if asked, and call them 'Boss' with respect.]\n"
        else:
            bot_context += "\n[NOTE: This is a regular user. Be polite, friendly, and helpful. Call them by their name.]\n"

        bot_context += "\nSubmission Rules: Tell users to click '🚀 SELL ID' -> Select Category -> Give Telegram username -> Upload Excel (.xlsx) file -> Give Payment number.\n"
        bot_context += f"Minimum IDs needed to submit: {SYSTEM_SETTINGS['min_id_limit']}\n"
        bot_context += f"Always start your response with 'আসসালামু আলাইকুম, {user_name}।'. Respond in friendly Bengali.\n"

        if CUSTOM_BOT_CONTEXT:
            bot_context += f"\n[Additional Admin Instructions/Rules]:\n{CUSTOM_BOT_CONTEXT}\n"

        max_retries = len(GROQ_API_KEYS)
        
        for attempt in range(max_retries):
            try:
                client = OpenAI(
                    api_key=GROQ_API_KEYS[CURRENT_API_INDEX],
                    base_url="https://api.groq.com/openai/v1", 
                )
                
                messages = [
                    {"role": "system", "content": bot_context}
                ]
                
                if image_data:
                    base64_image = base64.b64encode(image_data).decode('utf-8')
                    messages.append({
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                        ]
                    })
                    model_to_use = self.model_name_vision
                else:
                    messages.append({"role": "user", "content": prompt})
                    model_to_use = self.model_name_text

                response = client.chat.completions.create(
                    model=model_to_use,
                    messages=messages,
                    temperature=0.7,
                    max_tokens=1024
                )
                
                return response.choices[0].message.content.strip()
                
            except Exception as e:
                log.error(f"❌ Groq AI Error with API Index {CURRENT_API_INDEX}: {e}")
                CURRENT_API_INDEX = (CURRENT_API_INDEX + 1) % len(GROQ_API_KEYS)
                log.info(f"🔄 Switching to API Key Index: {CURRENT_API_INDEX}")
                time.sleep(1.5) 
                
        return f"আসসালামু আলাইকুম {user_name}, আমি এখন একটু ব্যস্ত বা সার্ভারে সমস্যা হচ্ছে। দয়া করে কিছুক্ষণ পর আবার চেষ্টা করুন।"

jarvis = JarvisAI()

def get_local_stats(text: str) -> str:
    text = text.lower()
    if any(w in text for w in ["সময়", "time", "বাজে"]):
        return f"⏰ স্যার, এখন সময় {datetime.now().strftime('%I:%M %p')}।"
    if any(w in text for w in ["battery", "চার্জ", "charge"]):
        try:
            res = subprocess.run(['termux-battery-status'], capture_output=True, text=True)
            data = json.loads(res.stdout)
            return f"🔋 বর্তমানে চার্জ আছে {data.get('percentage')}% এবং অবস্থা {data.get('status')}।"
        except: return "❌ ব্যাটারি তথ্য পেতে Termux:API অ্যাপ লাগবে।"
    if any(w in text for w in ["ram", "মেমোরি"]):
        try:
            res = subprocess.run(['free', '-m'], capture_output=True, text=True)
            return f"💾 সিস্টেম মেমোরি স্ট্যাটাস:\n`{res.stdout}`"
        except: return "❌ র‍্যাম তথ্য পাওয়া যায়নি।"
    if any(w in text for w in ["storage", "স্টোরেজ", "rom", "ডিস্ক", "disk"]):
        try:
            res = subprocess.run(['df', '-h'], capture_output=True, text=True)
            return f"💽 স্টোরেজ স্ট্যাটাস:\n`{res.stdout[:500]}`"
        except: return "❌ স্টোরেজ তথ্য পাওয়া যায়নি।"
    return None

# ════════════════════════════════════════════════════════════════
#  /start
# ════════════════════════════════════════════════════════════════
@bot.message_handler(commands=["start"])
def cmd_start(message):
    cid = message.chat.id
    if is_banned(cid): return

    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    register_user(message)
    name = message.from_user.first_name or "বন্ধু"

    welcome = (
        f"👋 আসসালামু আলাইকুম, *{name}*!\n\n"
        f"🔥 *{BOT_NAME}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{SYSTEM_SETTINGS['welcome_msg']}\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "🤖 যেকোনো প্রশ্নে *AI Chat* বাটন চাপুন!"
    )
    bot.send_message(cid, welcome, reply_markup=main_menu(cid), parse_mode="Markdown")

# ════════════════════════════════════════════════════════════════
#  MAIN TEXT HANDLER
# ════════════════════════════════════════════════════════════════
@bot.message_handler(func=lambda m: True, content_types=["text"])
def handle_text(message):
    cid  = message.chat.id
    
    if is_banned(cid): return
    
    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    register_user(message)
    text = (message.text or "").strip()

    if cid not in ai_sessions:
        ai_sessions[cid] = {"active": False}

    local_res = get_local_stats(text)
    if local_res:
        bot.send_message(cid, local_res, parse_mode="Markdown")
        return

    if text == "🚀 SELL ID 🚀":
        show_categories(cid)

    elif text == "📊 Price List":
        bot.send_message(cid, build_price_list(), parse_mode="Markdown")

    elif text == "🎧 Support":
        bot.send_message(cid, f"🟢 *সাপোর্টের জন্য যোগাযোগ করুন:*\n\n👤 {SYSTEM_SETTINGS['support_username']}\n\n_যেকোনো সমস্যায় সরাসরি মেসেজ করুন।_", parse_mode="Markdown")

    elif text == "📜 My History":
        show_history(cid)

    elif text == "👤 My Profile":
        show_profile(cid)

    elif text == "🏆 Top Sellers":
        show_leaderboard(cid)
        
    elif text == "🌐 Global Stats":
        show_global_stats(cid)
        
    elif text == "📩 Feedback":
        msg = bot.send_message(cid, "✍️ *আপনার অভিযোগ বা ফিডব্যাক লিখুন:*\n_(এটি সরাসরি অ্যাডমিনদের কাছে পাঠানো হবে)_", parse_mode="Markdown", reply_markup=cancel_btn())
        bot.register_next_step_handler(msg, step_user_feedback)

    # 🆕 Notice Board
    elif text == "📣 Notice Board":
        show_notice_board(cid)

    # 🆕 Check Submission Status
    elif text == "🔎 Check Status":
        msg = bot.send_message(cid, "🔎 *আপনার Submission ID দিন:*\n_উদাহরণ: SUB0001_", parse_mode="Markdown", reply_markup=cancel_btn())
        bot.register_next_step_handler(msg, step_check_status)

    # 🆕 Rules & FAQ
    elif text == "📚 Rules & FAQ":
        rules_text = (
            "📚 *বটের নিয়মাবলী এবং সাধারণ প্রশ্ন (FAQ)*\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "১. ফাইল অবশ্যই সঠিক ফরমেট অনুযায়ী *.xlsx* হতে হবে।\n"
            f"২. একসাথে মিনিমাম *{SYSTEM_SETTINGS['min_id_limit']}* টি আইডি সাবমিট করতে হবে।\n"
            "৩. অ্যাডমিন প্যানেল থেকে রিভিউ শেষে পেমেন্ট পাঠানো হয়, তাই সাবমিট করার পর দয়া করে অপেক্ষা করুন।\n"
            "৪. ইনভ্যালিড (ভুল) বা স্প্যাম আইডি সাবমিট করলে অ্যাকাউন্ট ব্যান করা হতে পারে।\n"
            f"৫. যেকোনো জিজ্ঞাসা বা পেমেন্ট সংক্রান্ত তথ্যের জন্য {SYSTEM_SETTINGS['support_username']} এ মেসেজ দিন।"
        )
        bot.send_message(cid, rules_text, parse_mode="Markdown")

    elif text == "🤖 AI Chat":
        if not SYSTEM_SETTINGS["ai_chat_enabled"] and not is_admin(cid):
            bot.send_message(cid, "🔴 বর্তমানে AI Chat সাময়িকভাবে বন্ধ আছে।")
            return
            
        ai_sessions[cid]["active"] = True
        bot.send_message(cid, "🤖 *AI Chat চালু হয়েছে!*\n\nবট সম্পর্কে যেকোনো কিছু জানতে পারেন।\nবন্ধ করতে *❌ Stop AI* বাটন চাপুন।", parse_mode="Markdown")

    elif text == "❌ Stop AI":
        ai_sessions[cid]["active"] = False
        bot.send_message(cid, "✅ AI Chat বন্ধ হয়েছে।", reply_markup=main_menu(cid), parse_mode="Markdown")

    elif text == "🔐 Admin Panel":
        if is_admin(cid):
            show_admin_panel(cid)

    elif ai_sessions[cid].get("active", False):
        user_name = message.from_user.first_name or "বন্ধু"
        handle_ai_message(cid, text, user_name)

    else:
        bot.send_message(cid, "❓ বুঝতে পারিনি। নিচের বাটনগুলো ব্যবহার করুন।", reply_markup=main_menu(cid))

def handle_ai_message(cid, text, user_name):
    bot.send_chat_action(cid, 'typing')
    try:
        past_memory = get_memory(cid)
        full_prompt = f"আগের মেমোরি: {past_memory}\nইউজার এখন বলেছে: {text}" if past_memory else text

        user_is_admin = is_admin(cid)
        reply = jarvis.get_response(full_prompt, user_name, is_admin_user=user_is_admin)
        
        save_memory(cid, f"User: {text}\nJarvis: {reply}")

        safe_reply = safe_md(reply)
        bot.send_message(cid, f"🤖 {safe_reply}", parse_mode="MarkdownV2")
    except Exception as e:
        log.error(f"Text handler error: {e}")
        bot.send_message(cid, "⚠️ টেক্সট প্রসেস করতে সমস্যা হয়েছে।")

# ════════════════════════════════════════════════════════════════
#  USER FEATURES (Original)
# ════════════════════════════════════════════════════════════════

def show_global_stats(cid):
    total_paid_tk = sum(s.get("total", 0.0) for s in all_submissions.values() if s.get("paid"))
    total_ids_bought = sum(s.get("live_qty", 0) for s in all_submissions.values() if s.get("paid"))
    
    stats_text = (
        "🌐 *Global Bot Statistics*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👥 মোট রেজিস্টার্ড ইউজার: *{len(registered_users)}* জন\n"
        f"📦 মোট আইডি কেনা হয়েছে: *{total_ids_bought}* টি\n"
        f"💰 মোট পেমেন্ট করা হয়েছে: *৳{total_paid_tk:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "_এটি আমাদের বটের লাইভ ট্রাস্ট স্কোর। আমরা দ্রুত পেমেন্টে বিশ্বাসী!_"
    )
    bot.send_message(cid, stats_text, parse_mode="Markdown")

def step_user_feedback(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    
    feedback_text = (message.text or "").strip()
    user_name = message.from_user.first_name or "Unknown"
    
    adm_cap = (
        f"📩 *New Feedback/Complain*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 From: {user_name} | `{cid}`\n"
        f"📝 Message:\n{feedback_text}"
    )
    
    try:
        bot.send_message(CHANNEL_ID, adm_cap, parse_mode="Markdown")
        bot.send_message(cid, "✅ *আপনার মেসেজ অ্যাডমিনদের কাছে সফলভাবে পাঠানো হয়েছে!*", parse_mode="Markdown", reply_markup=main_menu(cid))
    except Exception as e:
        bot.send_message(cid, "❌ মেসেজ পাঠাতে সমস্যা হয়েছে।", reply_markup=main_menu(cid))


def show_profile(cid):
    subs = user_submissions.get(cid, [])
    total_submitted = len(subs)
    total_approved = 0
    total_earned = 0.0

    for sid in subs:
        s = all_submissions.get(sid, {})
        if s.get("paid"):
            total_approved += s.get("live_qty", 0)
            total_earned += s.get("total", 0.0)

    # 🆕 Badge/Rank
    badge = get_user_badge(total_earned)

    profile_text = (
        "👤 *আপনার প্রোফাইল*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 User ID: `{cid}`\n"
        f"🎖️ Rank: *{badge}*\n"
        f"📦 মোট সাবমিট: *{total_submitted}* বার\n"
        f"✅ মোট এপ্রুভড আইডি: *{total_approved}* টি\n"
        f"💰 মোট আয়: *৳{total_earned:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        "🏅 *Rank Guide:*\n"
        "🌱 Starter → 🥉 Bronze (৳100+) → 🥈 Silver (৳500+) → 🥇 Gold (৳2000+) → 💎 Diamond (৳5000+)"
    )
    bot.send_message(cid, profile_text, parse_mode="Markdown")

def show_leaderboard(cid):
    user_earnings = {}
    for s in all_submissions.values():
        if s.get("paid"):
            u = s.get("user", "Unknown")
            user_earnings[u] = user_earnings.get(u, 0.0) + s.get("total", 0.0)

    if not user_earnings:
        bot.send_message(cid, "🏆 এখনো কেউ টপ সেলার তালিকায় নেই! প্রথমে সেল করে টপ সেলার হোন।")
        return

    sorted_users = sorted(user_earnings.items(), key=lambda x: x[1], reverse=True)[:5]
    
    lines = ["🏆 *Top 5 Sellers (লিডারবোর্ড)*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    medals = ["🥇", "🥈", "🥉", "🏅", "🏅"]
    for i, (user, amount) in enumerate(sorted_users):
        badge = get_user_badge(amount)
        lines.append(f"{medals[i]} {user}  —  *৳{amount:.2f}* {badge}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def export_submissions_to_excel(cid):
    if not all_submissions:
        bot.send_message(cid, "📭 এক্সপোর্ট করার মতো কোনো ডেটা নেই।")
        return
    
    bot.send_message(cid, "⏳ এক্সপোর্ট ফাইল তৈরি হচ্ছে, অপেক্ষা করুন...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions Data"
        headers = ["Sub ID", "User Name", "Category", "Rate", "Submitted Qty", "Live/Approved Qty", "Total Amount (Tk)", "Account Number", "Note", "Date", "Status"]
        ws.append(headers)
        
        for sid, s in all_submissions.items():
            status = "Paid" if s.get("paid") else "Pending"
            ws.append([
                sid, 
                s.get("user"), 
                s.get("type_name"), 
                s.get("rate"), 
                s.get("qty"), 
                s.get("live_qty", 0), 
                s.get("total", 0), 
                s.get("account"), 
                s.get("note"), 
                s.get("date"), 
                status
            ])
        
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        filename = f"Export_Data_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        bot.send_document(cid, document=(filename, file_stream), caption="📊 *All Submissions Exported Data*\n\nআপনার বটের সব ডাটা এখানে দেওয়া হলো।", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(cid, f"❌ এক্সপোর্ট করতে সমস্যা হয়েছে: {e}")

# ════════════════════════════════════════════════════════════════
#  🆕 NEW USER FEATURES
# ════════════════════════════════════════════════════════════════

def show_notice_board(cid):
    notice = SYSTEM_SETTINGS.get("notice_board", "")
    if not notice:
        bot.send_message(cid, "📣 *Notice Board*\n━━━━━━━━━━━━━━━━━━━━━━━━\n_এখনো কোনো নোটিশ নেই।_", parse_mode="Markdown")
    else:
        bot.send_message(cid, f"📣 *Notice Board*\n━━━━━━━━━━━━━━━━━━━━━━━━\n{notice}", parse_mode="Markdown")

def step_check_status(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    
    sub_id = (message.text or "").strip().upper()
    sub = all_submissions.get(sub_id)
    
    if not sub:
        bot.send_message(cid, f"❌ *'{sub_id}'* নামে কোনো Submission পাওয়া যায়নি।\nসঠিক Sub ID দিন (যেমন: SUB0001)", parse_mode="Markdown", reply_markup=main_menu(cid))
        return
    
    # Security: only allow owner or admin to check
    if sub.get("chat_id") != cid and not is_admin(cid):
        bot.send_message(cid, "🔒 এই Submission টি আপনার নয়।", reply_markup=main_menu(cid))
        return

    status_icon = "✅ *Paid/Completed*" if sub.get("paid") else "⏳ *Pending (Review বাকি)*"
    live = sub.get("live_qty", 0)
    fail = max(0, sub.get("qty", 0) - live) if sub.get("paid") else "-"
    total = sub.get("total", 0)

    status_text = (
        f"🔎 *Submission Status*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 Sub ID  : `{sub_id}`\n"
        f"📁 Type    : {sub.get('type_name')}\n"
        f"📅 Date    : {sub.get('date')}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Submit  : {sub.get('qty')} pcs\n"
        f"✅ Live    : {live} pcs\n"
        f"❌ Failed  : {fail} pcs\n"
        f"💰 Amount  : ৳{total:.2f}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📌 Status  : {status_icon}\n"
    )
    bot.send_message(cid, status_text, parse_mode="Markdown", reply_markup=main_menu(cid))

# ════════════════════════════════════════════════════════════════
#  PHOTO HANDLER
# ════════════════════════════════════════════════════════════════
@bot.message_handler(content_types=["photo"])
def handle_photo(message):
    cid = message.chat.id
    if is_banned(cid): return
    
    if SYSTEM_SETTINGS["maintenance_mode"] and not is_admin(cid):
        bot.send_message(cid, "🛠️ সিস্টেম বর্তমানে আপডেটের কাজে সাময়িকভাবে বন্ধ আছে। দয়া করে কিছুক্ষণ পর চেষ্টা করুন।")
        return

    user_name = message.from_user.first_name or "স্যার"
    
    if cid in user_data and user_data[cid].get("step") == "file":
        bot.send_message(cid, "❌ একটি *.xlsx* ফাইল পাঠান।", reply_markup=cancel_btn())
        return

    if cid in user_data and user_data[cid].get("step") == "review_screenshot":
        return

    bot.send_chat_action(cid, 'typing')
    try:
        file_info = bot.get_file(message.photo[-1].file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        caption = message.caption or "এই ছবিতে কী আছে?"
        user_is_admin = is_admin(cid)
        description = jarvis.get_response(caption, user_name, is_admin_user=user_is_admin, image_data=downloaded_file)
        safe_desc = safe_md(description)
        bot.reply_to(message, f"🤖 {safe_desc}", parse_mode="MarkdownV2")
    except Exception as e:
        log.error(f"Photo processing error: {e}")
        bot.send_message(cid, "❌ ছবিটি আমি পড়তে পারছি না বা মার্কডাউন সমস্যা হয়েছে।")

# ════════════════════════════════════════════════════════════════
#  SELL FLOW
# ════════════════════════════════════════════════════════════════
def show_categories(cid):
    markup = types.InlineKeyboardMarkup(row_width=1)
    for code, cat in CATEGORIES.items():
        if cat["status"] == "open":
            label = f"✅ {cat['name']}  —  {cat['rate']:.2f} Tk"
        else:
            label = f"🔴 {cat['name']}  [বন্ধ]"
        markup.add(types.InlineKeyboardButton(label, callback_data=f"sell_{code}"))

    bot.send_message(cid, "🛒 *ID বিক্রি করুন*\n━━━━━━━━━━━━━━━━━━━━━━━━\n\n👇 *ক্যাটাগরি সিলেক্ট করুন:*", reply_markup=markup, parse_mode="Markdown")

@bot.callback_query_handler(func=lambda c: c.data.startswith("sell_"))
def cb_sell(call):
    bot.answer_callback_query(call.id)
    cid  = call.message.chat.id
    code = call.data[5:]

    if code not in CATEGORIES: return
    cat = CATEGORIES[code]
    if cat["status"] != "open":
        bot.send_message(cid, "🔴 এই ক্যাটাগরি বর্তমানে বন্ধ আছে।")
        return

    user_data[cid] = {
        "type": code, "type_name": cat["name"],
        "rate": cat["rate"], "format": cat["format"],
        "date": datetime.now().strftime("%d %b %Y"),
        "step": "username"
    }

    msg = bot.send_message(
        cid,
        f"✅ *সিলেক্ট:* {cat['name']}\n💰 Rate: *{cat['rate']:.2f} Tk/pcs*\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n👤 *আপনার Telegram username দিন:*\n_উদাহরণ: @myusername_",
        reply_markup=cancel_btn(), parse_mode="Markdown"
    )
    bot.register_next_step_handler(msg, step_username)

def step_username(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    uname = (message.text or "").strip()
    if not re.match(r'^@[a-zA-Z0-9_]{4,32}$', uname):
        msg = bot.send_message(cid, "❌ *ভুল ফরম্যাট!*\n@ দিয়ে শুরু করুন\n_উদাহরণ: @myusername_", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_username)

    username_to_id[uname.lower()] = cid
    user_data[cid]["user"] = uname
    user_data[cid]["step"] = "file"

    fmt = user_data[cid]["format"]
    rate = user_data[cid]["rate"]

    msg = bot.send_message(
        cid,
        f"📤 *এখন Excel ফাইল পাঠান*\n━━━━━━━━━━━━━━━━━━━━━━━━\n📋 Column Layout:\n`{fmt}`\n\n"
        f"💰 Rate: *{rate:.2f} Tk/pcs*\n\n⚠️ শুধুমাত্র *.xlsx* ফাইল সাপোর্ট করে",
        reply_markup=cancel_btn(), parse_mode="Markdown"
    )
    bot.register_next_step_handler(msg, step_file_text)

def step_file_text(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if message.content_type == "document": return handle_docs(message)
    bot.send_message(cid, "❌ একটি *.xlsx* ফাইল পাঠান।", reply_markup=cancel_btn(), parse_mode="Markdown")
    bot.register_next_step_handler(message, step_file_text)

@bot.message_handler(content_types=["document"])
def handle_docs(message):
    cid = message.chat.id
    if is_banned(cid): return
    if cid not in user_data or user_data[cid].get("step") != "file": return

    fname = message.document.file_name or ""
    if not fname.lower().endswith(".xlsx"):
        msg = bot.reply_to(message, "❌ শুধু *.xlsx* ফাইল সাপোর্ট করে!\nএকটি Excel ফাইল পাঠান:", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_file_text)

    try:
        fi   = bot.get_file(message.document.file_id)
        raw  = bot.download_file(fi.file_path)
        wb   = openpyxl.load_workbook(filename=BytesIO(raw), data_only=True)
        ws   = wb.active
        qty  = sum(1 for row in ws.iter_rows(values_only=True) if any(c is not None and str(c).strip() for c in row))

        min_limit = SYSTEM_SETTINGS["min_id_limit"]
        if qty < min_limit:
            msg = bot.reply_to(message, f"❌ আপনি মাত্র *{qty}* টি আইডি দিয়েছেন।\n⚠️ মিনিমাম *{min_limit}* টি আইডি একসাথে দিতে হবে!\n\nসঠিক ফাইলটি আবার পাঠান:", reply_markup=cancel_btn(), parse_mode="Markdown")
            return bot.register_next_step_handler(msg, step_file_text)

        user_data[cid].update({"qty": qty, "file_name": fname, "file_id": message.document.file_id, "step": "number"})

        msg = bot.reply_to(
            message,
            f"✅ *ফাইল গৃহীত হয়েছে!*\n📊 মোট row: *{qty} pcs*\n\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "💳 *Payment Method: Bkash / Nagad / Rocket*\n\n📱 *আপনার payment নম্বর দিন:*\n_উদাহরণ: 01XXXXXXXXX_",
            reply_markup=cancel_btn(), parse_mode="Markdown"
        )
        bot.register_next_step_handler(msg, step_number)

    except Exception as e:
        log.error(f"File error: {e}")
        msg = bot.reply_to(message, f"❌ ফাইল পড়তে সমস্যা: `{e}`\nআবার চেষ্টা করুন:", reply_markup=cancel_btn(), parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_file_text)

def step_number(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    number = (message.text or "").strip()
    if not re.match(r'^01[0-9]{9}$', number):
        msg = bot.send_message(cid, "❌ *ভুল নম্বর!* ১১ ডিজিটের বাংলাদেশী নম্বর দিন\n_উদাহরণ: 01712345678_", reply_markup=cancel_btn(), parse_mode="Markdown")
        return bot.register_next_step_handler(msg, step_number)

    user_data[cid]["account"] = number
    user_data[cid]["step"]    = "note"

    msg = bot.send_message(cid, "📝 *কোনো নোট আছে?*\n_না থাকলে শুধু 'না' লিখুন_", reply_markup=cancel_btn(), parse_mode="Markdown")
    bot.register_next_step_handler(msg, step_note)

def step_note(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    if cid not in user_data: return

    note = (message.text or "").strip()
    finalize(cid, note)

def finalize(cid, note):
    global submission_count
    submission_count += 1
    sub_id = f"SUB{submission_count:04d}"
    d      = user_data[cid]
    est    = d["qty"] * d["rate"]

    all_submissions[sub_id] = {
        "chat_id":   cid,
        "user":      d["user"],
        "type_name": d["type_name"],
        "rate":      d["rate"],
        "qty":       d["qty"],
        "total":     0,
        "account":   d["account"],
        "note":      note,
        "date":      d["date"],
        "paid":      False,
        "live_qty":  0,
        "file_name": d["file_name"],
    }
    user_submissions.setdefault(cid, []).append(sub_id)

    receipt = (
        "✅ *সাবমিশন সম্পন্ন হয়েছে!*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 Sub ID  : `{sub_id}`\n📁 Type    : {d['type_name']}\n📅 Date    : {d['date']}\n📄 File    : `{d['file_name']}`\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📊 Total   : *{d['qty']} pcs*\n💰 Est.Amt : *{est:.2f} Tk* ({d['rate']} × {d['qty']})\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💳 Number  : `{d['account']}`\n📝 Note    : {note}\n\n"
        "⏳ *Admin review করার পর result পাঠাবে।*\n"
        f"📞 সমস্যায়: {SYSTEM_SETTINGS['support_username']}"
    )

    adm_cap = (
        f"📥 *New Submission #{sub_id}*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User    : {d['user']} | `{cid}`\n🏷️ Type    : {d['type_name']}\n📊 Rows    : *{d['qty']} pcs*\n"
        f"💰 Est.    : *{est:.2f} Tk*\n💳 Account : `{d['account']}`\n📝 Note    : {note}"
    )
    adm_btn = types.InlineKeyboardMarkup()
    adm_btn.add(types.InlineKeyboardButton("🟢 Review / Result দিন", callback_data=f"review_{sub_id}"))

    bot.send_message(cid, receipt, parse_mode="Markdown", reply_markup=main_menu(cid))
    try:
        bot.send_document(CHANNEL_ID, d["file_id"], caption=adm_cap, parse_mode="Markdown", reply_markup=adm_btn)
    except Exception as e: log.error(f"Channel send error: {e}")
    del user_data[cid]

# ════════════════════════════════════════════════════════════════
#  CALLBACKS & ADMIN CONTROLS
# ════════════════════════════════════════════════════════════════
@bot.callback_query_handler(func=lambda c: True)
def callback_handler(call):
    cid  = call.message.chat.id
    data = call.data

    if data == "cancel_flow":
        bot.answer_callback_query(call.id)
        user_data.pop(cid, None)
        bot.send_message(cid, "❌ বাতিল করা হয়েছে।", reply_markup=main_menu(cid))

    elif data.startswith("review_"):
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        sub_id = data[7:]
        sub = all_submissions.get(sub_id)
        if not sub: bot.send_message(cid, "❌ Submission পাওয়া যায়নি।"); return
        if sub["paid"]: bot.send_message(cid, "✅ এটি আগেই পেইড করা হয়েছে।"); return

        ch_msg_id = call.message.message_id if call.message.chat.id == CHANNEL_ID else None

        msg = bot.send_message(cid, f"🔖 *Review: #{sub_id}*\n👤 User: {sub['user']}\n📊 Total: {sub['qty']} pcs\n💳 Account: `{sub['account']}`\n\n✅ *কতটি Live/Success? (সংখ্যা দিন):*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_review, sub_id, ch_msg_id)

    elif data == "admin_pending_subs":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        
        pending_list = {k: v for k, v in all_submissions.items() if not v.get("paid")}
        if not pending_list:
            bot.send_message(cid, "✅ কোনো পেন্ডিং সাবমিশন নেই!")
            return
            
        m = types.InlineKeyboardMarkup(row_width=1)
        for sid, s in list(pending_list.items())[:20]: 
            m.add(types.InlineKeyboardButton(f"⏳ {sid} | {s['user']} | {s['qty']} pcs", callback_data=f"review_{sid}"))
        m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
        try: bot.edit_message_text("⏳ *পেন্ডিং সাবমিশন তালিকা:*\n_রিভিউ করতে নিচের সাবমিশনগুলোতে ক্লিক করুন_", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data == "already_done":
        bot.answer_callback_query(call.id, "✅ Already processed!")

    elif data == "admin_msg_user":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "👤 ইউজারের Telegram ID বা @username দিন:", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_msg_user_id)

    elif data == "admin_broadcast":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "📢 ব্রডকাস্ট মেসেজ লিখুন (বা ছবি + ক্যাপশন পাঠান):")
        bot.register_next_step_handler(msg, step_broadcast)

    elif data == "admin_refresh":
        bot.answer_callback_query(call.id, "🔄 Refreshed!")
        show_admin_panel(cid)

    elif data == "admin_all_subs":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        if not all_submissions: bot.send_message(cid, "📭 কোনো সাবমিশন নেই।"); return
        lines = ["📋 *সর্বশেষ ২০ সাবমিশন:*", "━━━━━━━━━━━━━━━━━━"]
        for sid, s in list(all_submissions.items())[-20:]:
            icon = "✅" if s.get("paid") else "⏳"
            lines.append(f"{icon} `{sid}` | {s['user']} | ৳{s.get('total',0):.2f}")
        bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

    elif data == "admin_change_rate_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        m = types.InlineKeyboardMarkup(row_width=1)
        for code, cat in CATEGORIES.items():
            m.add(types.InlineKeyboardButton(f"✏️ {cat['name']}  →  {cat['rate']} Tk", callback_data=f"setrate_{code}"))
        m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
        try: bot.edit_message_text("💰 *কোন ক্যাটাগরির রেট পরিবর্তন করবেন?*", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data.startswith("setrate_"):
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        code = data[8:]
        if code not in CATEGORIES: return
        msg = bot.send_message(cid, f"✏️ *{CATEGORIES[code]['name']}*\nনতুন rate লিখুন (যেমন: 15.50):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_set_rate, code)

    elif data == "admin_change_status_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        _show_status_menu(cid, call.message.message_id)

    elif data.startswith("changestatus_"):
        if not is_admin(cid): return
        code = data[13:]
        if code not in CATEGORIES: return
        cur = CATEGORIES[code]["status"]
        CATEGORIES[code]["status"] = "closed" if cur == "open" else "open"
        new = CATEGORIES[code]["status"]
        bot.answer_callback_query(call.id, f"{'✅ Open' if new=='open' else '🔴 Closed'}")
        _show_status_menu(cid, call.message.message_id)

    elif data == "admin_train_ai":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "🧠 *AI Training*\n\nজার্ভিসকে নতুন কী শেখাতে চান বা কোন নিয়ম যোগ করতে চান তা লিখে পাঠান:\n_(এটি তার আগের নিয়মের সাথে যুক্ত হবে।)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_train_ai)

    elif data == "admin_all_control":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        m = types.InlineKeyboardMarkup(row_width=1)
        m.add(types.InlineKeyboardButton(f"📉 Min ID Limit (Now: {SYSTEM_SETTINGS['min_id_limit']})", callback_data="admin_change_min_limit"))
        m.add(types.InlineKeyboardButton(f"🎧 Update Support User", callback_data="admin_change_support"))
        m.add(types.InlineKeyboardButton(f"📝 Edit Welcome Message", callback_data="admin_edit_welcome"))
        m.add(types.InlineKeyboardButton(f"📣 Set Notice Board", callback_data="admin_set_notice"))  # 🆕
        
        ai_status = "✅ ON" if SYSTEM_SETTINGS["ai_chat_enabled"] else "🔴 OFF"
        m.add(types.InlineKeyboardButton(f"🤖 Global AI Chat: {ai_status}", callback_data="admin_toggle_ai"))
        
        maint_status = "🔴 ON (Paused)" if SYSTEM_SETTINGS["maintenance_mode"] else "✅ OFF (Running)"
        m.add(types.InlineKeyboardButton(f"🛠️ Maintenance Mode: {maint_status}", callback_data="admin_toggle_maint"))
        
        m.add(types.InlineKeyboardButton("🚫 Ban / Unban User", callback_data="admin_ban_user_menu"))
        m.add(types.InlineKeyboardButton("💾 Backup Data (JSON)", callback_data="admin_db_backup"))
        m.add(types.InlineKeyboardButton("🧹 Clear All AI Memory", callback_data="admin_clear_ai_memory"))
        m.add(types.InlineKeyboardButton("⏰ Bot Uptime", callback_data="admin_uptime"))  # 🆕
        m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
        try:
            bot.edit_message_text("🎛️ *Wall / All Control Panel*\n\nসিস্টেমের গুরুত্বপূর্ণ লিমিট এবং সেটিং কন্ট্রোল করুন:", cid, call.message.message_id, reply_markup=m, parse_mode="Markdown")
        except: pass

    elif data == "admin_export_data":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        export_submissions_to_excel(cid)

    elif data == "admin_change_min_limit":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, f"📉 *বর্তমান মিনিমাম আইডি লিমিট:* {SYSTEM_SETTINGS['min_id_limit']} pcs\n\nনতুন মিনিমাম লিমিট কত দিতে চান? (সংখ্যায় লিখুন):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_change_min_limit)

    elif data == "admin_change_support":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, f"🎧 *বর্তমান সাপোর্ট ইউজার:* {SYSTEM_SETTINGS['support_username']}\n\nনতুন সাপোর্ট ইউজারনেম দিন (যেমন: @NewSupport):", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_change_support)

    elif data == "admin_edit_welcome":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "📝 *নতুন ওয়েলকাম মেসেজ লিখুন:*\n_(যে লেখাটি /start দিলে দেখাবে)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_edit_welcome)
        
    elif data == "admin_ban_user_menu":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "🚫 *ইউজারকে ব্যান বা আনব্যান করতে তার Telegram ID দিন:*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_ban_unban_user)

    elif data == "admin_db_backup":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        handle_db_backup(cid)

    elif data == "admin_toggle_ai":
        if not is_admin(cid): return
        SYSTEM_SETTINGS["ai_chat_enabled"] = not SYSTEM_SETTINGS["ai_chat_enabled"]
        bot.answer_callback_query(call.id, f"AI Chat is now {'ON' if SYSTEM_SETTINGS['ai_chat_enabled'] else 'OFF'}")
        call.data = "admin_all_control"
        callback_handler(call)

    elif data == "admin_toggle_maint":
        if not is_admin(cid): return
        SYSTEM_SETTINGS["maintenance_mode"] = not SYSTEM_SETTINGS["maintenance_mode"]
        bot.answer_callback_query(call.id, f"Maintenance is now {'ON' if SYSTEM_SETTINGS['maintenance_mode'] else 'OFF'}")
        call.data = "admin_all_control"
        callback_handler(call)

    elif data == "admin_clear_ai_memory":
        if not is_admin(cid): return
        local_ai_memory.clear()
        bot.answer_callback_query(call.id, "✅ All AI Memory Cleared!")

    # ════════════════════════════════════════════════════════════════
    #  🆕 NEW ADMIN CALLBACKS
    # ════════════════════════════════════════════════════════════════

    elif data == "admin_set_notice":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        current = SYSTEM_SETTINGS.get("notice_board", "_(ফাঁকা)_")
        msg = bot.send_message(cid, f"📣 *Notice Board আপডেট করুন*\n\nবর্তমান নোটিশ:\n{current}\n\nনতুন নোটিশ লিখুন:\n_(মুছে দিতে 'clear' লিখুন)_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_set_notice)

    elif data == "admin_uptime":
        if not is_admin(cid): return
        delta = datetime.now() - BOT_START_TIME
        hours, rem = divmod(int(delta.total_seconds()), 3600)
        minutes, seconds = divmod(rem, 60)
        uptime_text = f"⏰ *Bot Uptime*\n━━━━━━━━━━━━━━━━━━━━━━━━\n🟢 চালু আছে: *{hours}h {minutes}m {seconds}s*\n📅 চালু হয়েছিল: {BOT_START_TIME.strftime('%d %b %Y %I:%M %p')}"
        bot.answer_callback_query(call.id)
        bot.send_message(cid, uptime_text, parse_mode="Markdown")

    elif data == "admin_search_sub":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "🔎 *Sub ID দিয়ে সার্চ করুন:*\n_উদাহরণ: SUB0001_", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_admin_search_sub)

    elif data == "admin_revenue_summary":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        show_revenue_summary(cid)

    elif data == "admin_user_history":
        bot.answer_callback_query(call.id)
        if not is_admin(cid): return
        msg = bot.send_message(cid, "👤 *যে ইউজারের হিস্টোরি দেখতে চান তার Telegram ID দিন:*", parse_mode="Markdown")
        bot.register_next_step_handler(msg, step_admin_view_user_history)


def _show_status_menu(cid, msg_id):
    m = types.InlineKeyboardMarkup(row_width=1)
    for code, cat in CATEGORIES.items():
        icon = "✅" if cat["status"] == "open" else "🔴"
        m.add(types.InlineKeyboardButton(f"{icon} {cat['name']}", callback_data=f"changestatus_{code}"))
    m.add(types.InlineKeyboardButton("🔙 Back", callback_data="admin_refresh"))
    try: bot.edit_message_text("⚙️ *Category ON/OFF করুন:*\n_ক্লিক করলেই পরিবর্তন হবে_", cid, msg_id, reply_markup=m, parse_mode="Markdown")
    except: pass

# ════════════════════════════════════════════════════════════════
#  ADMIN REVIEW WITH SCREENSHOT 
# ════════════════════════════════════════════════════════════════
def step_review(message, sub_id, ch_msg_id):
    cid = message.chat.id
    if not (message.text or "").isdigit():
        msg = bot.send_message(cid, "❌ শুধু সংখ্যা দিন।")
        return bot.register_next_step_handler(msg, step_review, sub_id, ch_msg_id)

    live = int(message.text)
    sub  = all_submissions.get(sub_id)
    if not sub: return

    user_data[cid] = {
        "step": "review_screenshot",
        "review_sub_id": sub_id,
        "live_qty": live,
        "ch_msg_id": ch_msg_id
    }

    amount = live * sub["rate"]
    msg = bot.send_message(cid, f"✅ Live: *{live}* টি।\n💰 Amount: *৳{amount:.2f}*\n\n📸 *এবার পেমেন্টের স্ক্রিনশট পাঠান:*\n_(স্ক্রিনশট না দিতে চাইলে `skip` বা `না` লিখুন)_", parse_mode="Markdown")
    bot.register_next_step_handler(msg, step_review_screenshot)

def step_review_screenshot(message):
    cid = message.chat.id
    if cid not in user_data or "review_sub_id" not in user_data[cid]:
        return handle_text(message)

    d = user_data[cid]
    sub_id    = d["review_sub_id"]
    live      = d["live_qty"]
    ch_msg_id = d["ch_msg_id"]
    
    sub = all_submissions.get(sub_id)
    if not sub:
        del user_data[cid]
        return bot.send_message(cid, "❌ Submission পাওয়া যায়নি।")

    fail   = max(0, sub["qty"] - live)
    amount = live * sub["rate"]
    sub.update({"live_qty": live, "total": amount, "paid": True})

    result = (
        "📊 *আপনার Result এসেছে!*\n━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"🆔 Sub ID  : `{sub_id}`\n📁 Type    : {sub['type_name']}\n📅 Date    : {sub['date']}\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Total   : {sub['qty']} pcs\n✅ Success : *{live} pcs*\n❌ Failed  : *{fail} pcs*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💰 Amount  : *{amount:.2f} Tk*\n   _({sub['rate']} Tk × {live} pcs)_\n💳 Account : `{sub['account']}`\n\n"
        "✅ *Payment প্রক্রিয়া সম্পন্ন হয়েছে!* 🎉\n"
        f"❓ সমস্যায়: {SYSTEM_SETTINGS['support_username']}"
    )

    try: 
        if message.photo:
            bot.send_photo(sub["chat_id"], message.photo[-1].file_id, caption=result, parse_mode="Markdown")
        else:
            bot.send_message(sub["chat_id"], result, parse_mode="Markdown")
    except Exception as e: 
        log.warning(f"Result send failed: {e}")

    if ch_msg_id:
        done_btn = types.InlineKeyboardMarkup()
        done_btn.add(types.InlineKeyboardButton(f"✅ Completed — {live} Paid", callback_data="already_done"))
        try: bot.edit_message_reply_markup(CHANNEL_ID, ch_msg_id, reply_markup=done_btn)
        except: pass

    bot.send_message(cid, f"✅ *#{sub_id} Result পাঠানো হয়েছে!*\nAmount: ৳{amount:.2f}", parse_mode="Markdown")
    del user_data[cid]

# ════════════════════════════════════════════════════════════════
#  🆕 NEW ADMIN TOOLS IMPLEMENTATION
# ════════════════════════════════════════════════════════════════

def step_set_notice(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    text = (message.text or "").strip()
    if text.lower() == "clear":
        SYSTEM_SETTINGS["notice_board"] = ""
        bot.send_message(cid, "✅ *Notice Board মুছে দেওয়া হয়েছে!*", parse_mode="Markdown")
    else:
        SYSTEM_SETTINGS["notice_board"] = text
        bot.send_message(cid, "✅ *Notice Board আপডেট করা হয়েছে!*\nইউজাররা এখন এটি দেখতে পাবে।", parse_mode="Markdown")
    show_admin_panel(cid)

def step_admin_search_sub(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    sub_id = (message.text or "").strip().upper()
    sub = all_submissions.get(sub_id)
    if not sub:
        bot.send_message(cid, f"❌ *'{sub_id}'* পাওয়া যায়নি।", parse_mode="Markdown")
        return
    
    status = "✅ Paid" if sub.get("paid") else "⏳ Pending"
    text = (
        f"🔎 *Submission Details: {sub_id}*\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 User: {sub.get('user')} | `{sub.get('chat_id')}`\n"
        f"📁 Type: {sub.get('type_name')}\n"
        f"📅 Date: {sub.get('date')}\n"
        f"📦 Submit: {sub.get('qty')} pcs\n"
        f"✅ Live: {sub.get('live_qty', 0)} pcs\n"
        f"💰 Amount: ৳{sub.get('total', 0):.2f}\n"
        f"💳 Account: `{sub.get('account')}`\n"
        f"📝 Note: {sub.get('note')}\n"
        f"📌 Status: {status}"
    )
    
    m = types.InlineKeyboardMarkup()
    if not sub.get("paid"):
        m.add(types.InlineKeyboardButton("🔍 Review এখনই করুন", callback_data=f"review_{sub_id}"))
    bot.send_message(cid, text, parse_mode="Markdown", reply_markup=m)

def show_revenue_summary(cid):
    cat_revenue = {}
    for s in all_submissions.values():
        if s.get("paid"):
            t = s.get("type_name", "Unknown")
            cat_revenue[t] = cat_revenue.get(t, 0.0) + s.get("total", 0.0)
    
    if not cat_revenue:
        bot.send_message(cid, "📊 এখনো কোনো পেইড সাবমিশন নেই।")
        return
    
    total_rev = sum(cat_revenue.values())
    lines = ["📊 *Revenue Summary by Category*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    sorted_cats = sorted(cat_revenue.items(), key=lambda x: x[1], reverse=True)
    for cat_name, amount in sorted_cats:
        pct = (amount / total_rev * 100) if total_rev > 0 else 0
        lines.append(f"▪️ {cat_name}\n   💰 ৳{amount:.2f} ({pct:.1f}%)")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append(f"💵 *মোট আয়: ৳{total_rev:.2f}*")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def step_admin_view_user_history(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    
    target_input = (message.text or "").strip()
    target_id = None
    
    if target_input.lstrip("-").isdigit():
        target_id = int(target_input)
    elif target_input.lower() in username_to_id:
        target_id = username_to_id[target_input.lower()]
    
    if not target_id:
        bot.send_message(cid, "❌ ইউজার পাওয়া যায়নি।")
        return
    
    subs = user_submissions.get(target_id, [])
    if not subs:
        bot.send_message(cid, f"📭 এই ইউজারের (`{target_id}`) কোনো সাবমিশন নেই।", parse_mode="Markdown")
        return
    
    total_earned = 0.0
    lines = [f"📋 *User History: `{target_id}`*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for sid in reversed(subs[-20:]):
        s = all_submissions.get(sid, {})
        icon = "✅" if s.get("paid") else "⏳"
        amt = s.get("total", 0)
        total_earned += amt
        lines.append(f"{icon} `{sid}` | {s.get('type_name','?')} | ৳{amt:.2f}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━━━")
    lines.append(f"💰 মোট আয়: *৳{total_earned:.2f}* | 🎖️ {get_user_badge(total_earned)}")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

# ════════════════════════════════════════════════════════════════
#  ORIGINAL ADMIN TOOLS
# ════════════════════════════════════════════════════════════════

def step_edit_welcome(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    SYSTEM_SETTINGS["welcome_msg"] = message.text.strip()
    bot.send_message(cid, "✅ *ওয়েলকাম মেসেজ সফলভাবে আপডেট করা হয়েছে!*", parse_mode="Markdown")
    show_admin_panel(cid)

def step_ban_unban_user(message):
    cid = message.chat.id
    target_id = message.text.strip()
    if not target_id.isdigit():
        bot.send_message(cid, "❌ শুধু ইউজারের Telegram ID (সংখ্যায়) দিন।")
        return
    
    target_id = int(target_id)
    if target_id in ADMIN_IDS:
        bot.send_message(cid, "❌ অ্যাডমিনকে ব্যান করা যাবে না!")
        return

    if target_id in BANNED_USERS:
        BANNED_USERS.remove(target_id)
        bot.send_message(cid, f"✅ ইউজার `{target_id}` কে *আনব্যান* করা হয়েছে!", parse_mode="Markdown")
    else:
        BANNED_USERS.add(target_id)
        bot.send_message(cid, f"🚫 ইউজার `{target_id}` কে *ব্যান* করা হয়েছে!", parse_mode="Markdown")

def handle_db_backup(cid):
    bot.send_message(cid, "⏳ ডাটাবেস ব্যাকআপ তৈরি হচ্ছে...")
    try:
        db_dump = {
            "submissions": all_submissions,
            "user_submissions": user_submissions,
            "registered_users": list(registered_users),
            "banned_users": list(BANNED_USERS),
            "settings": SYSTEM_SETTINGS
        }
        json_data = json.dumps(db_dump, indent=4, default=str)
        file_stream = BytesIO(json_data.encode("utf-8"))
        
        filename = f"DB_Backup_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        bot.send_document(cid, document=(filename, file_stream), caption="💾 *সম্পূর্ণ ডাটাবেস ব্যাকআপ*\n\nআপনার বটের সব সেভড ডাটা এখানে আছে।", parse_mode="Markdown")
    except Exception as e:
        bot.send_message(cid, f"❌ ব্যাকআপ নিতে সমস্যা হয়েছে: {e}")

def step_train_ai(message):
    global CUSTOM_BOT_CONTEXT
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    CUSTOM_BOT_CONTEXT = message.text.strip()
    save_memory(cid, f"Admin Training: {CUSTOM_BOT_CONTEXT}")
    bot.send_message(cid, "✅ *AI সফলভাবে Train করা হয়েছে!*\nএখন থেকে জার্ভিস আপনার দেওয়া নতুন নিয়ম মেনে চলবে।", parse_mode="Markdown")
    show_admin_panel(cid)

def step_change_min_limit(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    try:
        new_limit = int(message.text.strip())
        SYSTEM_SETTINGS["min_id_limit"] = new_limit
        bot.send_message(cid, f"✅ *সফল!*\nএখন থেকে কেউ {new_limit} টির কম আইডি জমা দিতে পারবে না।", parse_mode="Markdown")
        show_admin_panel(cid)
    except ValueError:
        msg = bot.send_message(cid, "❌ ভুল ফরম্যাট! শুধু সংখ্যা দিন (যেমন: 10):")
        bot.register_next_step_handler(msg, step_change_min_limit)

def step_change_support(message):
    cid = message.chat.id
    if message.text in MENU_BUTTONS: return handle_text(message)
    SYSTEM_SETTINGS['support_username'] = message.text.strip()
    bot.send_message(cid, f"✅ *সফল!*\nসাপোর্ট ইউজার আপডেট করা হয়েছে: {SYSTEM_SETTINGS['support_username']}", parse_mode="Markdown")
    show_admin_panel(cid)

def step_msg_user_id(message):
    uid_input = (message.text or "").strip()
    target = None
    if uid_input.lstrip("-").isdigit(): target = int(uid_input)
    elif uid_input.lower() in username_to_id: target = username_to_id[uid_input.lower()]
    else: bot.send_message(message.chat.id, "❌ *পাওয়া যায়নি!*", parse_mode="Markdown"); return
    msg = bot.send_message(message.chat.id, "✍️ মেসেজ বা ছবি পাঠান:")
    bot.register_next_step_handler(msg, step_send_to_user, target)

def step_send_to_user(message, target):
    try:
        if message.photo: bot.send_photo(target, message.photo[-1].file_id, caption=message.caption or "")
        else: bot.send_message(target, f"📩 *Admin Message:*\n\n{message.text}", parse_mode="Markdown")
        bot.send_message(message.chat.id, f"✅ পাঠানো হয়েছে → `{target}`", parse_mode="Markdown")
    except Exception as e: bot.send_message(message.chat.id, f"❌ ব্যর্থ: {e}")

def step_broadcast(message):
    ok = fail = 0
    bot.send_message(message.chat.id, f"⏳ ব্রডকাস্ট শুরু...")
    for uid in list(registered_users):
        try:
            if message.photo: bot.send_photo(uid, message.photo[-1].file_id, caption=message.caption or "")
            else: bot.send_message(uid, message.text, parse_mode="Markdown")
            ok += 1
        except: fail += 1
        time.sleep(0.05)
    bot.send_message(message.chat.id, f"📢 *ব্রডকাস্ট সম্পন্ন!*\n✅ সফল: {ok}\n❌ ব্যর্থ: {fail}", parse_mode="Markdown")

def step_set_rate(message, code):
    try:
        new_rate = float(message.text.strip())
        CATEGORIES[code]["rate"] = new_rate
        bot.send_message(message.chat.id, f"✅ *{CATEGORIES[code]['name']}*\nনতুন rate: *{new_rate:.2f} Tk*", parse_mode="Markdown")
        show_admin_panel(message.chat.id)
    except ValueError:
        msg = bot.send_message(message.chat.id, "❌ সংখ্যা দিন:")
        bot.register_next_step_handler(msg, step_set_rate, code)

def show_history(cid):
    subs = user_submissions.get(cid, [])
    if not subs: bot.send_message(cid, "📭 এখনো কোনো সাবমিশন নেই।"); return
    lines = ["📜 *আপনার সাবমিশন হিস্টোরি*", "━━━━━━━━━━━━━━━━━━━━━━━━"]
    for sid in reversed(subs[-15:]):
        s    = all_submissions.get(sid, {})
        icon = "✅" if s.get("paid") else "⏳"
        amt  = s.get("total", 0)
        lines.append(f"{icon} `{sid}` | {s.get('type_name','?')} | ৳{amt:.2f} | {s.get('date','?')}")
    bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")

def show_admin_panel(cid):
    total  = len(all_submissions)
    paid   = sum(1 for s in all_submissions.values() if s.get("paid"))
    tk     = sum(s.get("total", 0) for s in all_submissions.values() if s.get("paid"))
    users  = len(registered_users)
    banned = len(BANNED_USERS)
    
    # অ্যাডমিন প্যানেলের বাটনগুলো সুন্দর করে ক্যাটাগরি অনুযায়ী সাজানো হয়েছে
    m = types.InlineKeyboardMarkup(row_width=2)
    m.add(types.InlineKeyboardButton("⏳ Pending", callback_data="admin_pending_subs"),
          types.InlineKeyboardButton("📋 All Subs", callback_data="admin_all_subs"))
    m.add(types.InlineKeyboardButton("💰 Rates", callback_data="admin_change_rate_menu"),
          types.InlineKeyboardButton("📊 Revenue", callback_data="admin_revenue_summary"))
    m.add(types.InlineKeyboardButton("👁️ User History", callback_data="admin_user_history"),
          types.InlineKeyboardButton("📩 Msg User", callback_data="admin_msg_user"))
    m.add(types.InlineKeyboardButton("📢 Broadcast", callback_data="admin_broadcast"),
          types.InlineKeyboardButton("⚙️ Categories", callback_data="admin_change_status_menu"))
    m.add(types.InlineKeyboardButton("🧠 Train AI", callback_data="admin_train_ai"),
          types.InlineKeyboardButton("🔎 Search ID", callback_data="admin_search_sub"))
    m.add(types.InlineKeyboardButton("🎛️ System Settings", callback_data="admin_all_control"),
          types.InlineKeyboardButton("📥 Backup", callback_data="admin_export_data"))
    m.add(types.InlineKeyboardButton("🔄 Refresh", callback_data="admin_refresh"))
    
    admin_text = (
        "👑 *ADVANCED ADMIN CONTROL*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 Total Users : *{users}*\n"
        f"🚫 Banned      : *{banned}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"📥 Submissions : *{total}*\n"
        f"✅ Paid        : *{paid}*\n"
        f"⏳ Pending     : *{total - paid}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💵 Total Paid  : *৳{tk:.2f}*\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━"
    )
    
    bot.send_message(
        cid, 
        admin_text, 
        reply_markup=m, parse_mode="Markdown"
    )

@bot.message_handler(func=lambda m: True, content_types=["video", "audio", "voice", "sticker"])
def handle_media(message):
    cid = message.chat.id
    if is_banned(cid): return
    if cid in user_data and user_data[cid].get("step") == "file":
        bot.send_message(cid, "❌ একটি *.xlsx* ফাইল পাঠান।", reply_markup=cancel_btn(), parse_mode="Markdown")
    else:
        bot.send_message(cid, "❓ শুধু নিচের বাটনগুলো ব্যবহার করুন।", reply_markup=main_menu(cid))

# ════════════════════════════════════════════════════════════════
#  RUN
# ════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    log.info(f"🚀 {BOT_NAME} starting...")
    while True:
        try:
            bot.infinity_polling(timeout=60, long_polling_timeout=30, skip_pending=True)
        except Exception as e:
            log.error(f"Polling crashed: {e}")
            time.sleep(5)
